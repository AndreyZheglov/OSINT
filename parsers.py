import csv
import openpyxl
from openpyxl import load_workbook
import xlsxwriter
import json
import re
import pandas as pd
from datetime import datetime, date
from sqlalchemy.orm import declarative_base
from sqlalchemy import (
    create_engine,
    Column,
    Integer,
    Sequence,
    String,
    DateTime,
)
"""
    "CREATE TABLE IF NOT EXISTS {db_name}.{table_name} ( \
                    #   id int(11) NOT NULL AUTO_INCREMENT, \
                    #   full_name varchar(50) NOT NULL, \
                    age varchar(50) NULL, \
                    #   mob_tel varchar(50) NULL, \
                    home_tel varchar(50) NULL, \
                    contact_tel varchar(50) NULL, \
                    #   address varchar(50) NULL, \
                    birth_date varchar(50) NULL, \
                    birth_address varchar(50) NULL, \
                    #   gender varchar(50) NULL, \
                    #   email varchar(50) NOT NULL,\
                    passport_id varchar(50) NOT NULL,\
                    passport_issued_by varchar(50) NOT NULL,\
                    passport_issue_date varchar(50) NOT NULL,\
                    tax_id  varchar(50) NOT NULL,\
                    source varchar(480) NOT NULL,\
                    PRIMARY KEY(id) \
                   )
    """

def log_event(event, filename="userlog.txt"):
    with open(filename, "a", encoding="utf-8") as log:
        log.write(f"{datetime.now()} --> {event}")


def refresh_logs():

    # retrieving data from log
    with open("userlog.txt", "r", encoding="utf-8") as file:
        log_data = file.readlines()
    for record in log_data:
        print(record, type(record))
        record_date = try_parsing_date(record[:10])
        delta = datetime.today()-record_date
        if delta.days > 5:
            log_data.pop(log_data.index(record))

    # rewriting bot.log
    with open("userlog.txt", "w", encoding="utf-8") as file:
        file.writelines(log_data)


# necessary variables for futher operations
csv_header = [
    "ПІБ",                  # 1
    "ДН",                   # 2
    "АДРЕСА",               # 3
    "МІСЦЕ РОБОТИ/СЛУЖБИ",  # 4
    "ПОСАДА/ЗВАННЯ",        # 5
    "ТЕЛ",                  # 6
    "ЕЛ.ПОШТА",             # 7
    "СОЦМЕРЕЖІ",            # 8
    "ПАСПОРТ",              # 9
    "СЕРІЯ",                # 10
    "ІД"                    # 11
]
header_orm = {
    "ПІБ": "fullname",                  # 1
    "ДН": "birth",                   # 2
    "АДРЕСА": "address",               # 3
    "МІСЦЕ РОБОТИ/СЛУЖБИ": "workplace",  # 4
    "ПОСАДА/ЗВАННЯ": "position",        # 5
    "ТЕЛ": "phone_number",                  # 6
    "ЕЛ.ПОШТА": "email",             # 7
    "СОЦМЕРЕЖІ": "social",            # 8
    "ПАСПОРТ": "passport",              # 9
    "СЕРІЯ": "series",                # 10
    "ІД": "id_number"                    # 11
}


AVAILABLE_FORMATS = ['.csv', '.xls/.xlsx', '.txt']
PHOTO_FORMATS = ['jpeg', 'jpg', 'png', 'bmp']
csv_path = __file__.replace('parsers.py', 'person_info.csv')
Base = declarative_base()


class User:

    def __init__(self, username: str, PIN: int, telegram_id: int, admin: bool):

        self.username = username
        self.PIN = PIN
        self.telegram_id = telegram_id
        self.admin = admin

    def __str__(self):
        return f'''Ім'я користувача:{self.username}
ПІН:{self.PIN}
Телеграм ІД:{self.telegram_id}                
Адмін:{self.admin}                                
'''

    def to_json(self):
        return self.__dict__


class Person(Base):
    __tablename__ = 'person_info'

    id = Column(Integer, Sequence('user_id_seq'), primary_key=True)
    fullname = Column(String(60), unique=True)
    birth = Column(String(12), nullable=True)
    address = Column(String(120), nullable=True)
    workplace = Column(String(100), nullable=True)
    position = Column(String(60), nullable=True)
    phone_number = Column(String(60), nullable=True)
    email = Column(String(100), nullable=True)
    social = Column(String(200), nullable=True)
    passport = Column(String(60), nullable=True)
    series = Column(String(60), nullable=True)
    id_number = Column(String(60), nullable=True)


    def __repr__(self):
        return f"<{self.__dict__}>"

    def to_list(self):
        return [
        self.fullname,
        self.birth,
        self.address,
        self.workplace,
        self.position,
        self.phone_number,
        self.email,
        self.social,
        self.passport,
        self.series,
        self.id_number
]


def user_init(data: dict):
    return User(data['username'], data['PIN'], data['telegram_id'], data['admin'])


def parse_users():
    filename = __file__.replace('parsers.py', 'allowed_users.json')
    with open(filename, "r", encoding="utf-8") as file:
        data = json.load(file)
        users = [user_init(fragment) for fragment in data]
    return users


def users_write(userslist: list):
    json_list = [user.to_json() for user in userslist]
    with open('allowed_users.json', "w", encoding="utf-8") as file:
        json.dump(json_list, file)


def write_csv(data: list, header=csv_header, path=csv_path):

    with open(path, 'w', encoding="utf-8", newline='') as file:
        writer = csv.writer(file)
        writer.writerow(header)
        writer.writerows(data)


def available_formats():
    out_string = ""
    for i in range(len(AVAILABLE_FORMATS)):
        out_string  += f"{i+1}. {AVAILABLE_FORMATS[i]}" + '\n'
    return out_string


def available_headers():
    out_string = ""
    for header in csv_header:
        out_string += f"{csv_header.index(header)+1}. {header}\n"
    out_string += "\nПримітка:\nДН: дата народження\nІД: ідентифікаційний номер"
    return out_string


def update_element(updatable_list: list, input_list: list):

    for index in range(len(updatable_list)):
        updatable_list[index] = input_list[index] if input_list[index] not in {"-", " ", "None"} \
            else updatable_list[index]


def find_element(data, header: list, key: str, value: str):

    output_arr = []
    if len(data) == 0:
        return None
    for sublist in data:
        if sublist[header.index(key)] in value:
            output_arr.append(sublist)

    return (output_arr, data.index(output_arr[0])) if len(output_arr) > 0 else (None, None)


def list_to_text(person: Person, header=csv_header):
    data = person.to_list()
    reply_text = f"ПІБ: {data[0]}\n"
    for index, header_index in zip(range(1, len(data)), range(1, len(csv_header))):
        text = f"{csv_header[header_index]}: {data[index]}\n" if data[index] is not None \
                                                                 and data[index] != 'None' else ""
        reply_text += text
    return reply_text


def get_info_by_parameter(key: str, value: str,  session):
    print(f"{key}:{value}")
    person = None
    search_value = f"%{value}%"
    if key == "ПІБ":
        person = session.query(Person).filter(Person.fullname.like(search_value)).all()
    elif key == "Адреса":
        person = session.query(Person).filter(Person.address.like(search_value)).all()
    elif key in "Місце роботи/служби":
        person = session.query(Person).filter(Person.workplace.like(search_value)).all()
    elif key in "Посада/Звання":
        person = session.query(Person).filter(Person.position.like(search_value)).all()
    elif key in "Телефон":
        person = session.query(Person).filter(Person.phone_number.like(search_value)).all()
    elif key in "ІНПП":
        person = session.query(Person).filter(Person.id_number.like(search_value)).all()
    else:
        return "Ключа не знайдено"

    if type(person) == list and len(person) > 1:
        return [element.to_list() for element in person]

    if type(person) == list and len(person) == 1:
        return list_to_text(person[0])

    elif type(person) == Person:
        return list_to_text(person)

    return "Інформації не знайдено."


def write_xlsx(filename: str, data: list, header=csv_header):
    workbook = xlsxwriter.Workbook(filename)
    worksheet = workbook.add_worksheet('Список осіб')
    worksheet.write_row(0, 0, header)
    for element in data:
        worksheet.write_row(data.index(element)+1, 0, element)
    workbook.close()


def try_parsing_date(text):
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%d.%m.%Y', '%d/%m/%Y', '%d.%m.%Y'):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    raise ValueError('no valid date format found')


def date_parse(data: str):
    """
    Formats of date can be found by the function:
        -    1900/12/01
        -    2019.01.25
        -    2099-10-30
        -    vice versa
    """
    result = re.search(r"(0[1-9]|[12][0-9]|3[01])[-/.](?:0[1-9]|1[012])[-/.](?:19\d{2}|20[01][0-9]|20[0-9]{2})",
                       data)
    return result.group() if result != None else None


def phone_parse(data: str):

    """
    Phone numbers can be found:
        -   +919367788755
        -   8989829304
        -   +16308520397
        -   786-307-3615
    """

    result = re.findall(r"[\+]?[(]?[0-9]{3}[)]?[-\s\.]?[0-9]{3}[-\s\.]?[0-9]{4,6}", data)
    return ",".join(result) if result != None else None


def email_parse(data: str):

    result = re.findall(r"(?:^|\s)[\w!#$%&'*+/=?^`{|}~-](\.?[\w!#$%&'*+/=?^`{|}~-]+)*@\w+[.-]?\w*\.[a-zA-Z]{2,3}\b",
                        data)
    return ",".join(result) if result != [] else None


def sites_parse(data: str):
    result = re.findall(r"http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+", data)
    return ",".join(result) if result != [] else None


# general parsing choice
def parser(session, filename: str, engine=None, delimiter="\t"):

    filetype = filename[len(filename)-5:len(filename)]

    if ".csv" in filetype:
        header, data = csv_parse(filename)
    elif ".txt" in filetype:
        header, data = txt_parse(filename)
    elif ".xlsx" in filetype:
        flag = xlsx_parse_with_pandas(filename, engine)
        return flag
    else:
        print(f"Parser for {filetype} has not been impemented yet.")
        return False

    flag = to_database(data, session, header)
    return flag


# db parser
def db_parse(database_name: str, table_name):
    try:
        engine = create_engine(f"sqlite:///")
        dataframe = pd.read_csv()
    except Exception:
        return -1
    pass


# csv parser
def csv_parse(filename: str):

    with open(filename, 'r', encoding="utf-8") as file:
        reader = csv.reader(file)
        header = next(reader)
        data = [row for row in reader]

    return header, data


# text parser
def txt_parse(filename: str, delimiter='\t'):

    with open(filename, 'r') as file:
        reader = file.read().splitlines()
        header = reader[0].split(delimiter)
        data = [element.split(delimiter) for element in reader[1:]]
    return header, data


# json_parser
# def json_parse(filename: str):
#     with open(filename, 'r') as file:
#         json_data = json.load(file)
#
#     pass


# xlsx parser
def xlsx_parse(filename: str):

    workbook = load_workbook(filename=filename)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        data = []
        for row in sheet.rows:
            row_list = [cell.value for cell in row if not isinstance(cell, openpyxl.cell.cell.MergedCell)]
            data.append(row_list)
        return data[0], data[1:]


# writing to csv
def to_csv(data: list, out_file=csv_path, header=csv_header):

    # loading data from csv
    path = __file__.replace('parsers.py', 'person_info.csv')
    with open(path, 'r', encoding='utf-8') as file:
        reader = csv.reader(file)
        header_1 = next(reader)
        info_from_csv = [element for element in reader]

    # set for further check
    checker_set = {"None", "-"}

    # standardizing header
    upper_header = [element.upper().strip() for element in header]

    # data standardization
    csv_data = []
    for element in data:
        csv_element = []
        for header_element in csv_header:
            header_element_upper = header_element.upper()
            if header_element_upper in upper_header:
                if header_element_upper == "ДН":
                    csv_element.append(try_parsing_date(element[upper_header.index(header_element_upper)]))
                else:
                    data_element = str(element[upper_header.index(header_element_upper)]).strip()
                    csv_element.append(data_element)
            else:
                csv_element.append(None)
        print(csv_element)

    # check for empty data
        if len(set(csv_element) - checker_set) > 1:
            if len(info_from_csv) > 0:
                updatable_list, index = find_element(info_from_csv, csv_header, "ПІБ", csv_element[0])
                if updatable_list != None:
                    update_element(updatable_list[0], csv_element)
                    info_from_csv[index] = updatable_list[0]
                    print(f"Record for {csv_element[0]} has been modified")
                    continue
            csv_data.append(csv_element)

    # forming a resulting data list
    info_from_csv.extend(csv_data)
    print(info_from_csv)

    # appending to csv
    write_csv(csv_data, csv_header)

    return True


def to_database(data, session, header=csv_header):
    print("Sending data to database...")
    # data standardize
    upper_header = [element.upper().strip() for element in header]

    # reorganizing data
    csv_data = []
    for element in data:
        csv_element = []
        for header_element in csv_header:
            header_element_upper = header_element.upper()
            if header_element_upper in upper_header:
                data_element = element[upper_header.index(header_element_upper)].strip() if type(element) == str \
                    else element[upper_header.index(header_element_upper)]
                csv_element.append(data_element)
            else:
                csv_element.append(None)

        if csv_element[0] == 'None':
            continue

        if csv_element not in csv_data:
            csv_data.append(csv_element)
        if len(csv_data) // 1000 == 0:
            print(f"Розмір даних: {len(csv_data)} записів")
    for element in csv_data:
        print(element)
        data = session.query(Person).filter().all()
        person_from_db = session.query(Person).filter_by(fullname=element[0]).first()
        session.commit()
        if not person_from_db and element[0] != None:
            person = Person(
                fullname=element[0],
                birth=element[1],
                address=element[2],
                workplace=element[3],
                position=element[4],
                phone_number=element[5],
                email=element[6],
                social=element[7],
                passport=element[8],
                series=element[9],
                id_number=element[10],
            )
            session.add(person)
            session.commit()
    return True


def xlsx_parse_with_pandas(filename: str, engine):

    # reading data from excel
    df = pd.read_excel(filename, header=0)

    # data equalization due to headers that are available
    for header in df.columns:
        header = header.upper()
    df_columns = set(df.columns)
    equal_headers = df_columns & set(csv_header)
    dropout_columns = list(df_columns - equal_headers)
    df = df.drop(columns=dropout_columns)
    for column in df.columns:
        df.rename(columns={column: header_orm[column]}, inplace=True)

    # dropping duplicates from dataframe
    df.drop_duplicates(subset="fullname")
    params = str([column for column in df.columns]).replace("[", "").replace("]", "").replace("'", "")

    if params:
        # writing data to database
        df.to_sql('temp_person_info', con=engine, if_exists='replace', chunksize=10000, index=False)
        print(params)
        # merging information from temporary table to an existing
        sql = f"""
                INSERT INTO person_info
                  ({params})
                SELECT DISTINCT *
                  FROM temp_person_info source
                  WHERE NOT EXISTS(select fullname from person_info)
                 GROUP BY source.fullname;
            """

        with engine.begin() as conn:  # TRANSACTION
            # conn.execute("DROP TABLE temp_person_info")
            conn.execute(sql)
            conn.execute("DROP TABLE temp_person_info")
        return True
    else:
        return False
